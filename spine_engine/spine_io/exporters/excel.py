######################################################################################################################
# Copyright (C) 2017-2021 Spine project consortium
# This file is part of Spine Engine.
# Spine Engine is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser General
# Public License as published by the Free Software Foundation, either version 3 of the License, or (at your option)
# any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
Framework for exporting a database to Excel file.

:author: P. Vennström (VTT), A. Soininen (VTT), M. Marin (KTH)
:date:   31.1.2020
"""

import numpy as np
from openpyxl import Workbook
from spinedb_api.export_functions import (
    export_expanded_parameter_values,
    export_object_groups,
    export_alternatives,
    export_scenarios,
    export_scenario_alternatives,
)

# FIXME: Use multiple sheets if data doesn't fit


def export_spine_database_to_xlsx(db_map, filepath):
    """Writes data from a Spine database into an excel file.

    Args:
        db_map (spinedb_api.DatabaseMapping): database mapping.
        filepath (str): destination path.
    """
    wb = Workbook()
    _write_parameter_values(db_map, wb)
    _write_object_groups(db_map, wb)
    _write_alternatives(db_map, wb)
    _write_scenarios(db_map, wb)
    _write_scenario_alternatives(db_map, wb)
    wb.save(filepath)
    wb.close()


def _write_parameter_values(db_map, wb):
    """Writes entity parameter values, one sheet per type-annotated entity class.

    Args:
        db_map (DatabaseMapping)
        wb (openpyxl.Workbook): excel workbook to write to
    """
    data, metadata = export_expanded_parameter_values(db_map)
    for key, class_data in data.items():
        ws = wb.create_sheet()
        ws.title = key
        class_metadata = metadata[key]
        class_metadata = dict(sheet_type="entity", **class_metadata)
        _write_metadata(ws, class_metadata)
        row_offset = len(class_metadata) + 2
        index_dim_count = class_metadata.get("index_dim_count")
        if index_dim_count:
            _write_indexed_values(ws, class_data, row_offset, index_dim_count)
        else:
            _write_data(ws, class_data, row_offset)


def _write_indexed_values(ws, data, row_offset, index_dim_count):
    header = data[0]
    index_start = next(iter(k for k, field in enumerate(header) if field == "index"))
    index_stop = index_start + index_dim_count
    keyed_data = {}  # (entity, alternative, parameter) -> index -> value
    indexes = set()
    for item in data[1:]:
        entity_alternative = tuple(item[:index_start])
        index = tuple(item[index_start:index_stop])
        indexes.add(index)
        for parameter, value in zip(header[index_stop:], item[index_stop:]):
            key = entity_alternative + (parameter,)
            if isinstance(value, float) and np.isnan(value):
                value = "nan"
            keyed_data.setdefault(key, {})[index] = value
    indexes = sorted(indexes)
    # Create columns
    columns = []
    # Vertical header, spans index_dim_count columns, (entity, alternative) are in the right-most one
    for j in range(index_dim_count - 1):
        column = ["-"] * index_start + ["index"] + [index[j] for index in indexes]
        columns.append(column)
    column = header[:index_start] + ["index"] + [index[-1] for index in indexes]
    columns.append(column)
    # Data
    for key, indexed_values in keyed_data.items():
        column = list(key) + [indexed_values.get(index) for index in indexes]
        columns.append(column)
    # Write columns
    for j, column in enumerate(columns):
        for i, value in enumerate(column):
            ws.cell(row=row_offset + i, column=j + 1).value = value


def _write_object_groups(db_map, wb):
    """Writes object groups, one sheet per object class.

    Args:
        db_map (DatabaseMapping)
        wb (openpyxl.Workbook): excel workbook to write to
    """
    groups_per_class = {}
    for class_name, group_name, member_name in export_object_groups(db_map):
        groups_per_class.setdefault(class_name, []).append((group_name, member_name))
    for class_name, data in groups_per_class.items():
        ws = wb.create_sheet()
        ws.title = f"{class_name}_group"
        metadata = {"sheet_type": "object group", "class_name": class_name}
        _write_metadata(ws, metadata)
        data.insert(0, ("group", "member"))
        _write_data(ws, data, len(metadata) + 2)


def _write_alternatives(db_map, wb):
    """Writes alternatives, one sheet.

    Args:
        db_map (DatabaseMapping)
        wb (openpyxl.Workbook): excel workbook to write to
    """
    data = export_alternatives(db_map)
    ws = wb.create_sheet()
    ws.title = "alternative"
    metadata = {"sheet_type": "alternative"}
    _write_metadata(ws, metadata)
    data.insert(0, ("alternative", "description"))
    _write_data(ws, data, len(metadata) + 2)


def _write_scenarios(db_map, wb):
    """Writes scenarios, one sheet.

    Args:
        db_map (DatabaseMapping)
        wb (openpyxl.Workbook): excel workbook to write to
    """
    data = export_scenarios(db_map)
    ws = wb.create_sheet()
    ws.title = "scenario"
    metadata = {"sheet_type": "scenario"}
    _write_metadata(ws, metadata)
    data.insert(0, ("scenario", "active", "description"))
    _write_data(ws, data, len(metadata) + 2)


def _write_scenario_alternatives(db_map, wb):
    """Writes scenario alternatives, one sheet.

    Args:
        db_map (DatabaseMapping)
        wb (openpyxl.Workbook): excel workbook to write to
    """
    data = export_scenario_alternatives(db_map)
    # Patch before alternative
    for i, row in enumerate(data):
        if row[-1] is None:
            data[i][-1] = "N/A"
    ws = wb.create_sheet()
    ws.title = "scenario_alternative"
    metadata = {"sheet_type": "scenario alternative"}
    _write_metadata(ws, metadata)
    data.insert(0, ("scenario", "alternative", "before alternative"))
    _write_data(ws, data, len(metadata) + 2)


def _write_metadata(ws, metadata):
    for k, (key, value) in enumerate(metadata.items()):
        ws.cell(row=k + 1, column=1).value = key
        ws.cell(row=k + 1, column=2).value = value


def _write_data(ws, data, row_offset):
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            ws.cell(row=row_offset + i, column=j + 1).value = value
