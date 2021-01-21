######################################################################################################################
# Copyright (C) 2017-2021 Spine project consortium
# This file is part of Spine Engine.
# Spine Engine is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser General
# Public License as published by the Free Software Foundation, either version 3 of the License, or (at your option)
# any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
Contains ExcelConnector class and a help function.

:author: P. Vennström (VTT)
:date:   1.6.2019
"""

from itertools import islice, takewhile
import io
from openpyxl import load_workbook
from spinedb_api import (
    RelationshipClassMapping,
    ObjectClassMapping,
    ObjectGroupMapping,
    AlternativeMapping,
    ScenarioMapping,
    ScenarioAlternativeMapping,
    from_database,
    ParameterValueFormatError,
)
from ..io_api import SourceConnection


class ExcelConnector(SourceConnection):
    """Template class to read data from another QThread."""

    # name of data source, ex: "Text/CSV"
    DISPLAY_NAME = "Excel"

    # dict with option specification for source.
    OPTIONS = {
        "header": {"type": bool, "label": "Has header", "default": False},
        "row": {"type": int, "label": "Skip rows", "Minimum": 0, "default": 0},
        "column": {"type": int, "label": "Skip columns", "Minimum": 0, "default": 0},
        "read_until_col": {"type": bool, "label": "Read until empty column on first row", "default": False},
        "read_until_row": {"type": bool, "label": "Read until empty row on first column", "default": False},
    }

    FILE_EXTENSIONS = "*.xlsx;;*.xlsm;;*.xltx;;*.xltm"

    def __init__(self, settings):
        super().__init__(settings)
        self._filename = None
        self._wb = None

    def connect_to_source(self, source):
        """saves filepath

        Arguments:
            source {str} -- filepath
        """
        if source:
            self._filename = source
            try:
                # FIXME: there seems to be no way of closing the workbook
                # when read_only=True, read file into memory first and then
                # open to avoid locking file while toolbox is running.
                with open(self._filename, "rb") as bin_file:
                    in_mem_file = io.BytesIO(bin_file.read())
                self._wb = load_workbook(in_mem_file, read_only=True)
            except Exception as error:
                raise error

    def disconnect(self):
        """Disconnect from connected source.
        """
        if self._wb:
            self._wb.close()
            self._wb = None
        self._filename = None

    def get_tables(self):
        """Method that should return Excel sheets as mappings and their options.

        Returns:
            dict: Sheets as mappings and options for each sheet or an empty dictionary if no workbook.

        Raises:
            Exception: If something goes wrong.
        """
        if not self._wb:
            return {}
        try:
            sheets = {}
            for sheet in self._wb.sheetnames:
                mapping, option = create_mappings_from_sheet(self._wb[sheet])
                sheets[sheet] = {"mapping": mapping, "options": option}
            return sheets
        except Exception as error:
            raise error

    def get_data_iterator(self, table, options, max_rows=-1):
        """
        Return data read from data source table in table. If max_rows is
        specified only that number of rows.
        """
        if not self._wb:
            return iter([]), [], 0

        if not table in self._wb:
            # table not found
            return iter([]), [], 0
        worksheet = self._wb[table]

        # get options
        has_header = options.get("header", False)
        skip_rows = options.get("row", 0)
        skip_columns = options.get("column", 0)
        stop_at_empty_col = options.get("read_until_col", False)
        stop_at_empty_row = options.get("read_until_row", False)

        if max_rows == -1:
            max_rows = None
        else:
            max_rows += skip_rows

        read_to_col = None
        try:
            first_row = next(islice(worksheet.iter_rows(), skip_rows, max_rows))
            if stop_at_empty_col:
                # find first empty col in top row and use that as a stop
                num_cols = 0
                for i, column in enumerate(islice(first_row, skip_columns, None)):

                    if column.value is None:
                        read_to_col = i + skip_columns
                        break
                    num_cols = num_cols + 1
            else:
                num_cols = len(first_row) - skip_columns
        except StopIteration:
            # no data
            num_cols = 0

        header = []
        rows = worksheet.iter_rows()
        rows = islice(rows, skip_rows, max_rows)
        # find header if it has one
        if has_header:
            try:
                header = [c.value for c in islice(next(rows), skip_columns, read_to_col)]
            except StopIteration:
                # no data
                return iter([]), [], 0

        # iterator for selected columns and and skipped rows
        data_iterator = (list(cell.value for cell in islice(row, skip_columns, read_to_col)) for row in rows)
        if stop_at_empty_row:
            # add condition to iterator
            condition = lambda row: row[0] is not None
            data_iterator = takewhile(condition, data_iterator)

        return data_iterator, header, num_cols

    def get_mapped_data(self, tables_mappings, options, table_types, table_row_types, max_rows=-1):
        """
        Overrides io_api method to check for some parameter_value types.
        """
        mapped_data, errors = super().get_mapped_data(tables_mappings, options, table_types, table_row_types, max_rows)
        value_pos = 3  # from (class, entity, parameter, value, *optionals)
        for key in ("object_parameter_values", "relationship_parameter_values"):
            for k, row in enumerate(mapped_data.get(key, [])):
                value = row[value_pos]
                if isinstance(value, str) and value.startswith("{") and value.endswith("}"):
                    try:
                        value = from_database(value)
                        row = row[:value_pos] + (value,) + row[value_pos + 1 :]
                        mapped_data[key][k] = row
                    except ParameterValueFormatError:
                        pass
        return mapped_data, errors


def get_mapped_data_from_xlsx(filepath):
    """Returns mapped data from given Excel file assuming it has the default Spine Excel format.

    Args:
        filepath (str): path to Excel file
    """
    connector = ExcelConnector(None)
    connector.connect_to_source(filepath)
    mappings_per_sheet = {}
    options_per_sheet = {}
    for sheet, data in connector.get_tables().items():
        if data["mapping"]:
            mappings_per_sheet[sheet] = data["mapping"]
        if data["options"]:
            options_per_sheet[sheet] = data["options"]
    types = row_types = dict.fromkeys(mappings_per_sheet, {})
    mapped_data, errors = connector.get_mapped_data(mappings_per_sheet, options_per_sheet, types, row_types)
    connector.disconnect()
    return mapped_data, errors


def create_mappings_from_sheet(ws):
    """
    Checks if sheet has the default Spine Excel format, if so creates a
    mapping object for each sheet.
    """
    metadata = _get_metadata(ws)
    if not metadata:
        return None, None
    header_row = len(metadata) + 2
    if metadata["sheet_type"] == "entity":
        index_dim_count = metadata.get("index_dim_count")
        header = _get_header(ws, header_row, index_dim_count)
        if not header:
            return None, None
        return _create_entity_mappings(metadata, header, index_dim_count)
    options = {"header": True, "row": len(metadata) + 1, "read_until_col": True, "read_until_row": True}
    if metadata["sheet_type"] == "object group":
        mapping = ObjectGroupMapping.from_dict({"name": metadata["class_name"], "groups": 0, "members": 1})
        return [mapping], options
    if metadata["sheet_type"] == "alternative":
        mapping = AlternativeMapping.from_dict({"name": 0})
        return [mapping], options
    if metadata["sheet_type"] == "scenario":
        mapping = ScenarioMapping.from_dict({"name": 0, "active": 1})
        return [mapping], options
    if metadata["sheet_type"] == "scenario alternative":
        mapping = ScenarioAlternativeMapping.from_dict(
            {"name": 0, "scenario_name": 0, "alternative_name": 1, "before_alternative_name": 2}
        )
        return [mapping], options
    return None, None


def _get_metadata(ws):
    metadata = {}
    for key, value in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if not key:
            break
        metadata[key] = value
    return metadata


def _get_header(ws, header_row, index_dim_count):
    if index_dim_count:
        # Vertical header
        header = []
        header_column = index_dim_count
        row = header_row
        while True:
            label = ws.cell(row=row, column=header_column).value
            if label == "index":
                break
            header.append(label)
            row += 1
        return header
    # Horizontal
    header = []
    column = 1
    while True:
        label = ws.cell(row=header_row, column=column).value
        if not label:
            break
        header.append(label)
        column += 1
    return header


def _create_entity_mappings(metadata, header, index_dim_count):
    class_name = metadata["class_name"]
    entity_type = metadata["entity_type"]
    map_dict = {"name": class_name}
    ent_alt_map_type = "row" if index_dim_count else "column"
    if entity_type == "object":
        map_dict["map_type"] = "ObjectClass"
        map_dict["objects"] = {"map_type": ent_alt_map_type, "reference": 0}
        mapping_class = ObjectClassMapping
    elif entity_type == "relationship":
        entity_dim_count = metadata["entity_dim_count"]
        map_dict["map_type"] = "RelationshipClass"
        map_dict["object_classes"] = header[:entity_dim_count]
        map_dict["objects"] = [{"map_type": ent_alt_map_type, "reference": i} for i in range(entity_dim_count)]
        mapping_class = RelationshipClassMapping
    else:
        return None, None
    value_type = metadata.get("value_type")
    if value_type is not None:
        value = {"value_type": value_type}
        if index_dim_count:
            value["extra_dimensions"] = list(range(index_dim_count))
        p_ref = len(header) if index_dim_count else -1
        map_dict["parameters"] = {
            "map_type": "ParameterValue",
            "name": {"map_type": "row", "reference": p_ref},
            "value": value,
            "alternative_name": {"map_type": ent_alt_map_type, "reference": header.index("alternative")},
        }
    mapping = mapping_class.from_dict(map_dict)
    options = {
        "header": not index_dim_count,
        "row": len(metadata) + 1,
        "read_until_col": True,
        "read_until_row": not index_dim_count,
    }
    return [mapping], options
